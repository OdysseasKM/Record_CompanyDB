import openpyxl
import random
import sqlite3
from datetime import datetime, timedelta

source = "data and db/artist-song-album.xlsx"
start = datetime(2000, 1, 1)
end = datetime(2022, 1, 1)

def open_excel_file():
    global wb,sheet
    # open files
    wb = openpyxl.load_workbook(source) 
    sheet = wb["Sheet1"]

def open_db():
    global db,cursor
    db = sqlite3.connect('record-company.db')
    cursor = db.cursor()

def create_artist_ids():
    i=99
    cur_artist_name = " "
    for rowNum in range (2,sheet.max_row):
        artist_name = sheet.cell(row=rowNum, column=3).value
        
        if (cur_artist_name==artist_name):
            sheet.cell(row=rowNum, column=2).value = i
        else:
            i+=1
            sheet.cell(row=rowNum, column=2).value = i
            cur_artist_name = artist_name
        # print(artist_name)
    wb.save(source)
            
def artist_data():
    ids = []

    exists = False
    for rowNum in range (2,sheet.max_row):
        #find artist id from each comumn
        artist_id = sheet.cell(row=rowNum, column=2).value

        #check if id has already added in file
        for i in range(1,len(ids)+1):
            # check first from the end of the pinax
            if (ids[-i]==artist_id): 
                exists = True
                break
        
        if (exists==False):
            #new aritst
            nickname = sheet.cell(row=rowNum, column=3).value
            r = random.randint(1,35)
            country = sheet.cell(row=r, column = 5).value
            #print(str(artist_id)+ " "+ nickname+ " "+ country)
            sql = """INSERT INTO ARTIST
                        VALUES(?, ?, ?);"""
            cursor.execute(sql,(artist_id, nickname, country))   
            ids.append(artist_id)
        exists = False

def add_format(release_id,format_id,vinyl,cd,digital):
    sql = """INSERT INTO FORMAT VALUES(?, ?);"""
    cursor.execute(sql,(release_id,format_id))
     
    if (vinyl==True):
        sales = random.randint(10,10000)
        cost = random.random() * 100
        cost = round(cost,2)
        sql = """INSERT INTO VINYL VALUES(?, ?, ?);"""
        cursor.execute(sql,(format_id,sales,cost))
       
    if (cd==True):
        sales = random.randint(10,10000)
        cost = random.random() * 100
        cost = round(cost,2)
        sql = """INSERT INTO CD VALUES(?, ?, ?);"""
        cursor.execute(sql,(format_id,sales,cost))

    if (digital==True):
        views = random.randint(100,1000000) 
        sql = """INSERT INTO DIGITAL VALUES(?, ?);"""
        cursor.execute(sql,(format_id,views))

def add_rating_for_song(song_id):
    for j in range(random.randint(1,5)):
        rate = random.randint(1,5)
        sql = """INSERT INTO RATING (stars,song_id) VALUES(?, ?);"""
        cursor.execute(sql,(rate,song_id))

def add_rating_for_video(video_id):
    for j in range(random.randint(1,5)):
        rate = random.randint(1,5)
        sql = """INSERT INTO RATING (stars,video_id) VALUES(?, ?);"""
        cursor.execute(sql,(rate,video_id))

def release_data():
    song_id=100
    release_id=100
    format_id=100
    languages = ["greek","english","spanish"]
    current_album = ""
    for rowNum in range (2,sheet.max_row):
        column2 = sheet.cell(row=rowNum, column=1).value
        if (column2=="single"):
            #add release
            artist_id = sheet.cell(row=rowNum, column=2).value
            single_name = sheet.cell(row=rowNum, column=4).value
            random_date = start + timedelta(days=random.randint(0, (end - start).days))
            genre_id = random.randint(0,11)

            sql = """INSERT INTO RELEASE
                    VALUES(?, ?, ?, ?, ?);"""
            cursor.execute(sql,(release_id, artist_id, random_date, single_name, genre_id))
            
            # add_single
            duration = random.randint(100,500)
            studio_id = random.randint(0,4)
            language = random.choice(languages)

            sql = """INSERT INTO SONG (song_id,rel_id,duration,studio_id,lyrics_language,title)
                    VALUES(?, ?, ?, ?, ?, ?);"""
            cursor.execute(sql,(song_id, release_id,duration, studio_id, language, single_name))
            add_rating_for_song(song_id)
            add_format(release_id,format_id,True,True,True)
            add_contributes_in_table_for_song(release_id)
            if (random.randint(1,4)==1):f_ar = add_feature_in(artist_id,release_id)
            format_id+=1
            release_id +=1
            
            # video
            has_video = random.randint(0,10)
            # nine out of ten singles have video
            if (has_video!=0):
                # single has a video
                # add video release
                random_date = start + timedelta(days=random.randint(0, (end - start).days))
                sql = """INSERT INTO RELEASE
                    VALUES(?, ?, ?, ?, ?);"""
                cursor.execute(sql,(release_id, artist_id, random_date, single_name+" (video clip)", genre_id))

                duration = random.randint(100,500)
                sql = """INSERT INTO VIDEO
                    VALUES(?, ?, ?);"""
                cursor.execute(sql,(release_id, song_id, duration))
                add_rating_for_video(release_id)
                add_format(release_id,format_id,False,False,True)
                add_contributes_in_table_for_video(release_id)
                try:
                    add_feature_in2(f_ar,release_id)
                except Exception:pass
                format_id+=1
                release_id +=1
            song_id+=1

        else:
            # it is an album
            if (column2!=current_album):
                # add album as a release
                artist_id = sheet.cell(row=rowNum, column=2).value
                random_date = start + timedelta(days=random.randint(0, (end - start).days))
                genre_id = random.randint(0,11)

                sql = """INSERT INTO RELEASE
                        VALUES(?, ?, ?, ?, ?);"""
                cursor.execute(sql,(release_id, artist_id, random_date, column2, genre_id))
                
                sql = """INSERT INTO ALBUM
                        VALUES(?);"""
                cursor.execute(sql,(release_id,))
                add_format(release_id,format_id,True,True,True)
                # feature artists
                if (random.randint(1,3)==1):f_ar1=add_feature_in(artist_id,release_id)
                if (random.randint(1,3)==1):f_ar2=add_feature_in(artist_id,release_id)
                format_id+=1
                
                album_id = release_id
                release_id +=1
                current_album = column2
                language = random.choice(languages)

            # add song 
            duration = random.randint(100,500)
            studio_id = random.randint(0,4)
            song_title = sheet.cell(row=rowNum, column=4).value
            sql = """INSERT INTO SONG (song_id, album_id, duration,studio_id,lyrics_language,title)
                    VALUES(?, ?, ?, ?, ?, ?);"""
            cursor.execute(sql,(song_id,album_id,duration,studio_id,language,song_title))
            add_rating_for_song(song_id)

            # video
            has_video = random.randint(0,5)
            # one out of five songs have video clip
            if (has_video==0):
                # single has a video
                # add video release
                random_date = start + timedelta(days=random.randint(0, (end - start).days))
                sql = """INSERT INTO RELEASE
                    VALUES(?, ?, ?, ?, ?);"""
                cursor.execute(sql,(release_id, artist_id, random_date, song_title+" (video clip)", genre_id))
                add_contributes_in_table_for_video(release_id)

                duration = random.randint(100,500)
                sql = """INSERT INTO VIDEO
                    VALUES(?, ?, ?);"""
                cursor.execute(sql,(release_id, song_id, duration))
                add_rating_for_video(release_id)
                add_format(release_id,format_id,False,False,True)
                try :
                    if (random.randint(1,2)==1):add_feature_in2(f_ar1,release_id)
                    if (random.randint(1,2)==1):add_feature_in2(f_ar2,release_id)
                except Exception:pass
                format_id+=1
                release_id +=1
            song_id+=1

def read_data_from_sql(source):
    with open(source, 'r') as file:
        # Read all the lines
        lines = file.readlines()
        # Print the lines
        for line in lines:
            if (line=='\n'):continue
            else:
                cursor.execute(line)

def load_sql_code(source):
    with open(source, 'r') as file:
        code = file.read()
        cursor.executescript(code)

def add_contributes_in_table_for_song(rel_id):
    sheet2 = wb["Sheet2"]
    roles = ["songwritter","music producer","music director","composer"]
    for role in roles:
        r_ssn = random.randint(1,29)
        ssn = sheet2.cell(row=r_ssn, column=1).value
        sql = """INSERT INTO CONTRIBUTES_IN (contributor_id,rel_id,role) VALUES(?, ?, ?);"""
        cursor.execute(sql,(ssn, rel_id, role))

def add_contributes_in_table_for_video(rel_id):
    sheet2 = wb["Sheet2"]
    roles = ["video producer","director"]
    for role in roles:
        r_ssn = random.randint(1,29)
        ssn = sheet2.cell(row=r_ssn, column=1).value
        sql = """INSERT INTO CONTRIBUTES_IN (contributor_id,rel_id,role) VALUES(?, ?, ?);"""
        cursor.execute(sql,(ssn, rel_id, role))

def select_random_artist():
    return random.randint(100,106)

# random aritst
def add_feature_in(artist_id,release_id):
    while (True):
        feature_ar_id = select_random_artist()
        if (feature_ar_id!=artist_id):break
    sql = """INSERT INTO FEATURE_IN VALUES(?, ?);"""
    cursor.execute(sql,(artist_id, release_id))
    return feature_ar_id

# specific artist
def add_feature_in2(f_artist,release_id):
    sql = """INSERT INTO FEATURE_IN VALUES(?, ?);"""
    cursor.execute(sql,(f_artist), release_id)

def main():
    open_excel_file()
    open_db()
    # create_artist_ids()

    create_db = "data and db/record_company_sqlite_create.sql"
    add_data = "data and db/add_data_in_db.sql"

    load_sql_code(create_db)
    load_sql_code(add_data)

    artist_data()
    release_data()
    
    print("your datasets are ready for queries.")
    db.commit()
    wb.close()

if __name__ == "__main__":main()
            
		
