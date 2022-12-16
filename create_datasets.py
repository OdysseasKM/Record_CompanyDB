import openpyxl
import random
import sqlite3
from datetime import date

#paths
source = "artist-song-album.xlsx"

# open files
wb = openpyxl.load_workbook(source)
sheet = wb["Sheet1"]


db = sqlite3.connect('database.db')
cursor = db.cursor()

def create_artist_ids():
    prev_id = -1
    i=100
    for rowNum in range (2,sheet.max_row):
        #find artist id from each comumn
        artist_id = sheet.cell(row=rowNum, column=4).value

        if (prev_id == int(artist_id)):
            sheet.cell(row=rowNum, column=4).value = i
        else:
            sheet.cell(row=rowNum, column=4).value = i
            prev_id = artist_id
            i+=1
    wb.save(source)
            
def artist_data():
    ids = []

    exists = False
    for rowNum in range (2,sheet.max_row):
        #find artist id from each comumn
        artist_id = sheet.cell(row=rowNum, column=4).value

        #check if id has already added in file
        for i in range(1,len(ids)+1):
            # check first from the end of the pinax
            if (ids[-i]==artist_id): 
                exists = True
                break
        
        if (exists==False):
            #new aritst
            nickname = sheet.cell(row=rowNum, column=5).value
            r = random.randint(1,186)
            country = sheet.cell(row=r, column = 10).value
            #print(str(artist_id)+ " "+ nickname+ " "+ country)
            sql = """INSERT INTO ARTIST
                        VALUES(?, ?, ?);"""
            cursor.execute(sql,(artist_id, nickname, country))
            db.commit()
            ids.append(artist_id)
        exists = False

def release_data():
    today = date.today()
    for rowNum in range (2,sheet.max_row):
        artist_id = sheet.cell(row=rowNum, column=4).value
        release_id = sheet.cell(row=rowNum, column = 1).value
        release_title = sheet.cell(row=rowNum, column = 9).value
        sql = """INSERT INTO RELEASE
                VALUES(?, ?, ?, ?);"""
        cursor.execute(sql,(artist_id, release_id, today, release_title))
        db.commit()

        

def main():
    create_artist_ids()
    artist_data()
    release_data()
    print("your datasets are ready for queries.")
    wb.close()

if __name__ == "__main__":main()
            
		
